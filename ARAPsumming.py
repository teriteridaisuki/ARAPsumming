import openpyxl
import os
#定义汇总表的名字和工作表的名字
MainWorkBook_name = "债权债务表汇总模板"
WorkSheet_name_AR = "应收账款改"
WorkSheet_name_AP = "应付账款改"
CurrentARrow=2
CurrentAProw=2
MainWorkBook = openpyxl.load_workbook(MainWorkBook_name+".xlsx")
MainWorkSheet_AR=MainWorkBook[WorkSheet_name_AR]
MainWorkSheet_AP=MainWorkBook[WorkSheet_name_AP]
def excelsumming(filename):
    global CurrentAProw,CurrentARrow
    SubWorkBook=openpyxl.load_workbook(filename,data_only=True)
    SubAR=SubWorkBook[WorkSheet_name_AR]
    SubAP = SubWorkBook[WorkSheet_name_AP]
    for row in range(2,SubAR.max_row+1):
        if (SubAR.cell(row, 2).value==''or SubAR.cell(row, 2).value==None):
            continue
        for col in list(range(2, 16))+[29]:
            MainWorkSheet_AR.cell(CurrentARrow,col).value=SubAR.cell(row,col).value
        MainWorkSheet_AR.cell(CurrentARrow, 1).value = CurrentARrow-1
        MainWorkSheet_AR.cell(CurrentARrow, 2).value = int(SubAR.cell(row, 2).value)
        CurrentARrow+=1
    for row in range(2,SubAP.max_row+1):
        if (SubAP.cell(row, 2).value==''or SubAP.cell(row, 2).value==None):
            continue
        for col in range(3, 12):
            MainWorkSheet_AP.cell(CurrentAProw,col).value=SubAP.cell(row,col).value
        MainWorkSheet_AP.cell(CurrentAProw, 1).value = CurrentAProw-1
        MainWorkSheet_AP.cell(CurrentAProw, 2).value = int(SubAP.cell(row,2).value)
        CurrentAProw+=1
    MainWorkBook.save(MainWorkBook_name+"完成.xlsx")

def main():#遍历文件夹“文件源”里的文件汇总
    filenames=os.listdir(r"数据源")
    for filename in filenames:
        if filename[-5:]==".xlsx":#确保只汇总xlsx文件
            excelsumming(r"数据源/%s"%filename)
            print(filename)
main()